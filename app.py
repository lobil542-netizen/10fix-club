from flask import Flask, request, jsonify, render_template, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import uuid

app = Flask(__name__)
app.secret_key = "10fix_secret_2026"

EXCEL_FILE = "customers.xlsx"
PURCHASES_FILE = "purchases.xlsx"
VISITS_FILE = "visits.xlsx"
ADMIN_PASSWORD = "10fix2026"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "לקוחות"
        ws.append(["תאריך", "שם מלא", "עיר", "טלפון", "אימייל", "אישור שיווק", "מזהה"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        wb.save(EXCEL_FILE)

def init_purchases():
    if not os.path.exists(PURCHASES_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "רכישות"
        ws.append(["תאריך", "מזהה לקוח", "שם לקוח", "פריט", "מחיר", "הערה"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        wb.save(PURCHASES_FILE)

def init_visits():
    if not os.path.exists(VISITS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "כניסות"
        ws.append(["תאריך", "מזהה לקוח", "שם לקוח", "טלפון"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        wb.save(VISITS_FILE)

def log_visit(uid, name, phone):
    init_visits()
    wb = load_workbook(VISITS_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime("%d/%m/%Y %H:%M"), uid, name, phone])
    wb.save(VISITS_FILE)

def get_recent_visits(limit=20):
    init_visits()
    wb = load_workbook(VISITS_FILE)
    ws = wb.active
    visits = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            visits.append({
                "date": row[0] or "",
                "uid":  row[1] or "",
                "name": row[2] or "",
                "phone": row[3] or ""
            })
    visits.sort(key=lambda x: x["date"], reverse=True)
    return visits[:limit]

def migrate_uids():
    """תן מזהה לכל לקוח שאין לו"""
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    changed = False
    for row in ws.iter_rows(min_row=2):
        if row[1].value and (len(row) < 7 or not row[6].value):
            row[6].value = str(uuid.uuid4())[:8]
            changed = True
    if changed:
        wb.save(EXCEL_FILE)

def get_customers():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            customers.append({
                "date":    row[0] or "",
                "name":    row[1] or "",
                "city":    row[2] or "",
                "phone":   row[3] or "",
                "email":   row[4] or "",
                "consent": row[5] or "",
                "uid":     row[6] or ""
            })
    customers.sort(key=lambda x: x["name"])
    return customers

def get_customer_by_uid(uid):
    for c in get_customers():
        if c["uid"] == uid:
            return c
    return None

def get_purchases(uid=None):
    init_purchases()
    wb = load_workbook(PURCHASES_FILE)
    ws = wb.active
    purchases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            p = {
                "date":     row[0] or "",
                "uid":      row[1] or "",
                "name":     row[2] or "",
                "item":     row[3] or "",
                "price":    row[4] or "",
                "note":     row[5] or ""
            }
            if uid is None or p["uid"] == uid:
                purchases.append(p)
    purchases.sort(key=lambda x: x["date"], reverse=True)
    return purchases

def get_purchase_stats():
    purchases = get_purchases()
    stats = {}
    for p in purchases:
        uid = p["uid"]
        if uid not in stats:
            stats[uid] = {"count": 0, "last": ""}
        stats[uid]["count"] += 1
        if not stats[uid]["last"] or p["date"] > stats[uid]["last"]:
            stats[uid]["last"] = p["date"]
    return stats

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    data = request.json
    if not data:
        return jsonify({"error": "no data"}), 400

    full_name = data.get("full_name", "").strip()
    city      = data.get("city", "").strip()
    phone     = data.get("phone", "").strip()
    email     = data.get("email", "").strip()
    consent   = "כן" if data.get("marketing_consent") else "לא"

    if not full_name or not city or not phone:
        return jsonify({"error": "missing fields"}), 400

    customer_uid = str(uuid.uuid4())[:8]

    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        full_name, city, phone, email, consent, customer_uid
    ])
    wb.save(EXCEL_FILE)

    print(f"[+] לקוח חדש: {full_name} | {city} | {phone} | {customer_uid}")
    return jsonify({"status": "ok", "uid": customer_uid})

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect(url_for("admin"))
        return render_template("admin_login.html", error="סיסמה שגויה")

    if not session.get("admin"):
        return render_template("admin_login.html", error=None)

    migrate_uids()
    customers = get_customers()
    stats = get_purchase_stats()
    return render_template("admin.html", customers=customers, total=len(customers), stats=stats)

@app.route("/admin/customer/<uid>", methods=["GET", "POST"])
def customer_card(uid):
    if not session.get("admin"):
        return redirect(url_for("admin"))
    customer = get_customer_by_uid(uid)
    if not customer:
        return "לקוח לא נמצא", 404

    if request.method == "POST":
        item  = request.form.get("item", "").strip()
        price = request.form.get("price", "").strip()
        note  = request.form.get("note", "").strip()
        if item:
            init_purchases()
            wb = load_workbook(PURCHASES_FILE)
            ws = wb.active
            ws.append([
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                uid, customer["name"], item, price, note
            ])
            wb.save(PURCHASES_FILE)
        return redirect(url_for("customer_card", uid=uid))

    purchases = get_purchases(uid)
    total_spent = sum(float(p["price"]) for p in purchases if p["price"] and str(p["price"]).replace('.','').isdigit())
    return render_template("customer_card.html", customer=customer, purchases=purchases, total_spent=total_spent)

@app.route("/checkin", methods=["GET", "POST"])
def checkin():
    if request.method == "GET":
        return render_template("checkin.html")

    phone = request.form.get("phone", "").strip().replace("-", "").replace(" ", "")
    if not phone:
        return render_template("checkin.html", error="נא להזין מספר טלפון")

    # חיפוש לפי טלפון
    customers = get_customers()
    found = None
    for c in customers:
        c_phone = str(c["phone"]).replace("-", "").replace(" ", "")
        if c_phone == phone:
            found = c
            break

    if found:
        log_visit(found["uid"], found["name"], found["phone"])
        return render_template("checkin.html", success=True, name=found["name"])
    else:
        return render_template("checkin.html", not_found=True)

@app.route("/api/recent-checkins")
def api_recent_checkins():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(get_recent_visits(10))

@app.route("/terms")
def terms():
    return render_template("terms.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin"))

if __name__ == "__main__":
    init_excel()
    migrate_uids()
    init_purchases()
    init_visits()
    print("=" * 40)
    print("10FIX מועדון לקוחות - שרת פעיל")
    print("כתובת: http://localhost:5000")
    print("=" * 40)
    app.run(host="0.0.0.0", port=5000, debug=False)
